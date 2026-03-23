"""
AWP-073 – MCP Search Extension
Erweitert den MCP-Client um PDF/Office-Dokumenten-Suche.

Unterstützte Formate:
  - PDF  (via pypdf / pdfplumber)
  - DOCX (via python-docx)
  - XLSX (via openpyxl)
  - TXT / MD (direkt)

Registriert als MCP-Tool "search_documents".
Python 3.12 | AsyncIO
"""

from __future__ import annotations

import asyncio
import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

log = logging.getLogger("jarvis.mcp_search_ext")

# Allowed search roots (no traversal outside)
ALLOWED_ROOTS: list[Path] = [
    Path(__file__).parent.parent / "docs",
    Path(__file__).parent.parent / "strategy",
    Path(__file__).parent.parent / "logs",
]

MAX_RESULTS    = 10
SNIPPET_LEN    = 300


@dataclass
class DocHit:
    file: str
    page: int          # 0 = N/A
    snippet: str
    score: float       # simple keyword match count


@dataclass
class SearchResult:
    query: str
    hits: list[DocHit] = field(default_factory=list)
    total_scanned: int = 0
    error: str = ""


# ── Extractors ─────────────────────────────────────────────────────────────

def _extract_pdf(path: Path) -> list[tuple[int, str]]:
    """Returns list of (page_number, text) pairs."""
    try:
        import pdfplumber  # type: ignore
        pages = []
        with pdfplumber.open(path) as pdf:
            for i, page in enumerate(pdf.pages, 1):
                text = page.extract_text() or ""
                pages.append((i, text))
        return pages
    except ImportError:
        pass

    try:
        from pypdf import PdfReader  # type: ignore
        reader = PdfReader(str(path))
        return [
            (i + 1, page.extract_text() or "")
            for i, page in enumerate(reader.pages)
        ]
    except ImportError:
        log.warning("No PDF library available (install pdfplumber or pypdf)")
        return []


def _extract_docx(path: Path) -> list[tuple[int, str]]:
    try:
        from docx import Document  # type: ignore
        doc = Document(str(path))
        text = "\n".join(p.text for p in doc.paragraphs)
        return [(0, text)]
    except ImportError:
        log.warning("python-docx not installed")
        return []


def _extract_xlsx(path: Path) -> list[tuple[int, str]]:
    try:
        from openpyxl import load_workbook  # type: ignore
        wb = load_workbook(str(path), read_only=True, data_only=True)
        parts = []
        for sheet in wb.worksheets:
            rows = []
            for row in sheet.iter_rows(values_only=True):
                rows.append(" | ".join(str(c) for c in row if c is not None))
            parts.append((0, "\n".join(rows)))
        return parts
    except ImportError:
        log.warning("openpyxl not installed")
        return []


def _extract_text(path: Path) -> list[tuple[int, str]]:
    try:
        return [(0, path.read_text(encoding="utf-8", errors="replace"))]
    except OSError:
        return []


def _extract(path: Path) -> list[tuple[int, str]]:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return _extract_pdf(path)
    if suffix == ".docx":
        return _extract_docx(path)
    if suffix in (".xlsx", ".xlsm"):
        return _extract_xlsx(path)
    if suffix in (".txt", ".md", ".rst", ".log"):
        return _extract_text(path)
    return []


# ── Path guard ─────────────────────────────────────────────────────────────

def _is_allowed(path: Path) -> bool:
    resolved = path.resolve()
    return any(
        resolved.is_relative_to(root.resolve())
        for root in ALLOWED_ROOTS
    )


# ── Search logic ───────────────────────────────────────────────────────────

def _score(text: str, terms: list[str]) -> float:
    """Simple keyword frequency score (case-insensitive)."""
    text_lower = text.lower()
    return sum(text_lower.count(t) for t in terms) / max(len(text), 1) * 1000


def _snippet(text: str, terms: list[str]) -> str:
    """Extract snippet around first term match."""
    text_lower = text.lower()
    for term in terms:
        idx = text_lower.find(term)
        if idx >= 0:
            start = max(0, idx - 80)
            end   = min(len(text), idx + SNIPPET_LEN)
            return "..." + text[start:end].replace("\n", " ") + "..."
    return text[:SNIPPET_LEN].replace("\n", " ")


def _search_sync(query: str, roots: list[Path]) -> SearchResult:
    terms = [t.lower() for t in re.split(r"\s+", query.strip()) if t]
    result = SearchResult(query=query)
    hits: list[DocHit] = []

    extensions = {".pdf", ".docx", ".xlsx", ".xlsm", ".txt", ".md", ".rst", ".log"}

    for root in roots:
        if not root.exists():
            continue
        for path in root.rglob("*"):
            if path.suffix.lower() not in extensions:
                continue
            if not _is_allowed(path):
                continue
            result.total_scanned += 1
            pages = _extract(path)
            for page_num, text in pages:
                if not text.strip():
                    continue
                score = _score(text, terms)
                if score > 0:
                    hits.append(DocHit(
                        file=str(path),
                        page=page_num,
                        snippet=_snippet(text, terms),
                        score=score,
                    ))

    hits.sort(key=lambda h: -h.score)
    result.hits = hits[:MAX_RESULTS]
    return result


# ── Async public API ───────────────────────────────────────────────────────

async def search_documents(
    query: str,
    roots: list[str] | None = None,
) -> SearchResult:
    """
    Search PDF/Office/text documents for query terms.

    Args:
        query: Space-separated search terms
        roots: Optional list of root directories (must be within allowed roots)
    """
    if not query.strip():
        return SearchResult(query=query, error="Empty query")

    search_roots = ALLOWED_ROOTS
    if roots:
        # Validate custom roots
        custom = [Path(r) for r in roots]
        search_roots = [r for r in custom if _is_allowed(r)]
        if not search_roots:
            return SearchResult(
                query=query,
                error="No allowed search roots provided",
            )

    loop = asyncio.get_event_loop()
    result = await loop.run_in_executor(None, _search_sync, query, search_roots)
    log.info(
        "Doc search '%s': %d hits / %d scanned",
        query, len(result.hits), result.total_scanned,
    )
    return result


# ── MCP Tool Registration ──────────────────────────────────────────────────

MCP_TOOL_SPEC = {
    "name": "search_documents",
    "description": "Search PDF, DOCX, XLSX, TXT and MD files for keywords",
    "inputSchema": {
        "type": "object",
        "properties": {
            "query": {
                "type": "string",
                "description": "Search terms (space-separated)",
            },
            "roots": {
                "type": "array",
                "items": {"type": "string"},
                "description": "Optional: limit search to these directories",
            },
        },
        "required": ["query"],
    },
}


async def handle_mcp_call(arguments: dict) -> list[dict]:
    """MCP tool handler — returns content list."""
    query = arguments.get("query", "")
    roots = arguments.get("roots")
    result = await search_documents(query, roots)

    if result.error:
        return [{"type": "text", "text": f"Error: {result.error}"}]

    if not result.hits:
        return [{"type": "text", "text": f"No results for '{query}' in {result.total_scanned} files."}]

    lines = [f"**{len(result.hits)} results** for `{query}` (scanned {result.total_scanned} files)\n"]
    for i, hit in enumerate(result.hits, 1):
        loc = f" (page {hit.page})" if hit.page else ""
        lines.append(f"**{i}. {Path(hit.file).name}**{loc}")
        lines.append(f"> {hit.snippet}")
        lines.append("")

    return [{"type": "text", "text": "\n".join(lines)}]
